"""
Excel Automation Script for Risk & Policy Analytics Dashboard
Author: Data Analyst
Date: April 2025
Purpose: Automated Excel reports with advanced formulas for client presentations
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime, timedelta
import warnings
warnings.filterwarnings('ignore')

class ExcelAutomation:
    def __init__(self, data_path):
        """Initialize Excel automation"""
        self.data_path = data_path
        self.df = None
        self.processed_df = None
        self.workbook = None
        
    def load_and_prepare_data(self):
        """Load and prepare data for Excel reports"""
        print("Loading data for Excel automation...")
        self.df = pd.read_csv(self.data_path)
        
        # Clean and prepare data
        df = self.df.copy()
        
        # Handle missing values
        df['Age'] = df['Age'].fillna(df['Age'].median())
        df['Annual Income'] = df['Annual Income'].fillna(df['Annual Income'].median())
        df['Number of Dependents'] = df['Number of Dependents'].fillna(0)
        df['Health Score'] = df['Health Score'].fillna(df['Health Score'].median())
        df['Previous Claims'] = df['Previous Claims'].fillna(0)
        df['Credit Score'] = df['Credit Score'].fillna(df['Credit Score'].median())
        df['Customer Feedback'] = df['Customer Feedback'].fillna('Average')
        
        # Convert dates
        df['Policy Start Date'] = pd.to_datetime(df['Policy Start Date'])
        
        # Create calculated columns
        df['Age Group'] = pd.cut(df['Age'], bins=[0, 25, 35, 45, 55, 65, 100], 
                                labels=['18-25', '26-35', '36-45', '46-55', '56-65', '65+'])
        
        df['Income Group'] = pd.cut(df['Annual Income'], bins=[0, 30000, 60000, 100000, 200000, float('inf')], 
                                   labels=['Low', 'Lower-Mid', 'Mid', 'Upper-Mid', 'High'])
        
        # Risk calculation
        df['Risk Score'] = self._calculate_risk_score(df)
        df['Risk Category'] = df['Risk Score'].apply(self._categorize_risk)
        
        # Premium metrics
        df['Premium Per Year'] = df['Premium Amount'] / df['Insurance Duration']
        df['Loss Ratio'] = (df['Previous Claims'] * 1000) / df['Premium Amount']
        df['Loss Ratio'] = df['Loss Ratio'].fillna(0)
        
        # Customer value
        df['Customer Value'] = df['Premium Amount'] * df['Insurance Duration']
        
        self.processed_df = df
        return df
    
    def _calculate_risk_score(self, df):
        """Calculate risk score for each policy"""
        risk_scores = []
        
        for _, row in df.iterrows():
            score = 0
            
            # Age factor
            if row['Age'] < 25 or row['Age'] > 65:
                score += 2
            elif 25 <= row['Age'] <= 35:
                score += 1
            
            # Claims history
            if row['Previous Claims'] > 2:
                score += 3
            elif row['Previous Claims'] > 0:
                score += 1
            
            # Health score
            if row['Health Score'] < 20:
                score += 2
            elif row['Health Score'] < 40:
                score += 1
            
            # Credit score
            if row['Credit Score'] < 500:
                score += 2
            elif row['Credit Score'] < 650:
                score += 1
            
            # Smoking status
            if row['Smoking Status'] == 'Yes':
                score += 2
            
            # Exercise frequency
            if row['Exercise Frequency'] == 'Rarely':
                score += 1
            
            risk_scores.append(score)
        
        return risk_scores
    
    def _categorize_risk(self, score):
        """Categorize risk based on score"""
        if score <= 2:
            return 'Low'
        elif score <= 4:
            return 'Medium'
        elif score <= 6:
            return 'High'
        else:
            return 'Very High'
    
    def create_executive_summary_report(self):
        """Create executive summary report with KPIs"""
        print("Creating executive summary report...")
        
        if self.processed_df is None:
            self.load_and_prepare_data()
        
        df = self.processed_df
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Executive Summary"
        
        # Set up headers and styling
        self._setup_worksheet_style(ws)
        
        # Title
        ws['A1'] = "RISK & POLICY ANALYTICS DASHBOARD"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:F1')
        
        ws['A2'] = f"Executive Summary Report - {datetime.now().strftime('%B %Y')}"
        ws['A2'].font = Font(size=12, italic=True)
        ws.merge_cells('A2:F2')
        
        # Key Performance Indicators
        ws['A4'] = "KEY PERFORMANCE INDICATORS"
        ws['A4'].font = Font(size=14, bold=True)
        ws.merge_cells('A4:F4')
        
        # Calculate KPIs
        total_policies = len(df)
        total_premiums = df['Premium Amount'].sum()
        avg_premium = df['Premium Amount'].mean()
        total_claims = df['Previous Claims'].sum()
        overall_loss_ratio = (df['Previous Claims'] * 1000).sum() / df['Premium Amount'].sum()
        high_risk_policies = len(df[df['Risk Category'].isin(['High', 'Very High'])])
        high_risk_pct = (high_risk_policies / total_policies) * 100
        
        # KPI Table
        kpi_data = [
            ['Metric', 'Value', 'Target', 'Status'],
            ['Total Policies', total_policies, 'N/A', 'Current'],
            ['Total Premium Volume', f'${total_premiums:,.2f}', 'N/A', 'Current'],
            ['Average Premium', f'${avg_premium:,.2f}', 'N/A', 'Current'],
            ['Total Claims', total_claims, 'N/A', 'Current'],
            ['Overall Loss Ratio', f'{overall_loss_ratio:.2%}', '<70%', 'Good' if overall_loss_ratio < 0.7 else 'Needs Attention'],
            ['High Risk Policies', f'{high_risk_pct:.1f}%', '<20%', 'Good' if high_risk_pct < 20 else 'Needs Attention']
        ]
        
        # Add KPI data to worksheet
        for row_idx, row_data in enumerate(kpi_data, start=6):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 6:  # Header row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
        
        # Create KPI table
        kpi_table = Table(displayName="KPITable", ref=f"A6:D{6+len(kpi_data)-1}")
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        kpi_table.tableStyleInfo = style
        ws.add_table(kpi_table)
        
        # Policy Performance Summary
        ws['A15'] = "POLICY PERFORMANCE SUMMARY"
        ws['A15'].font = Font(size=14, bold=True)
        ws.merge_cells('A15:F15')
        
        # Policy type performance
        policy_performance = df.groupby('Policy Type').agg({
            'Premium Amount': ['count', 'sum', 'mean'],
            'Previous Claims': 'sum',
            'Loss Ratio': 'mean'
        }).round(2)
        
        policy_performance.columns = ['Policy Count', 'Total Premiums', 'Avg Premium', 'Total Claims', 'Avg Loss Ratio']
        policy_performance = policy_performance.reset_index()
        
        # Add policy performance data
        headers = ['Policy Type', 'Policy Count', 'Total Premiums', 'Avg Premium', 'Total Claims', 'Avg Loss Ratio']
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=17, column=col_idx, value=header).font = Font(bold=True)
        
        for row_idx, (_, row) in enumerate(policy_performance.iterrows(), start=18):
            for col_idx, value in enumerate([row['Policy Type'], row['Policy Count'], 
                                          f"${row['Total Premiums']:,.2f}", f"${row['Avg Premium']:,.2f}", 
                                          row['Total Claims'], f"{row['Avg Loss Ratio']:.2%}"], start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Risk Analysis
        ws['A25'] = "RISK ANALYSIS"
        ws['A25'].font = Font(size=14, bold=True)
        ws.merge_cells('A25:F25')
        
        risk_analysis = df['Risk Category'].value_counts()
        risk_analysis_pct = df['Risk Category'].value_counts(normalize=True) * 100
        
        # Add risk analysis data
        risk_headers = ['Risk Category', 'Policy Count', 'Percentage']
        for col_idx, header in enumerate(risk_headers, start=1):
            ws.cell(row=27, column=col_idx, value=header).font = Font(bold=True)
        
        for row_idx, (category, count) in enumerate(risk_analysis.items(), start=28):
            percentage = risk_analysis_pct[category]
            ws.cell(row=row_idx, column=1, value=category)
            ws.cell(row=row_idx, column=2, value=count)
            ws.cell(row=row_idx, column=3, value=f"{percentage:.1f}%")
        
        # Add charts
        self._add_executive_charts(ws, df)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        return wb
    
    def _setup_worksheet_style(self, ws):
        """Set up worksheet styling"""
        # Define styles
        header_font = Font(size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply styles to all cells
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def _add_executive_charts(self, ws, df):
        """Add charts to executive summary"""
        # Premium by Policy Type Chart
        policy_premiums = df.groupby('Policy Type')['Premium Amount'].sum()
        
        chart1 = BarChart()
        chart1.title = "Total Premiums by Policy Type"
        chart1.x_axis.title = "Policy Type"
        chart1.y_axis.title = "Premium Amount ($)"
        
        data = Reference(ws, min_col=2, min_row=17, max_row=17+len(policy_premiums), max_col=3)
        cats = Reference(ws, min_col=1, min_row=18, max_row=17+len(policy_premiums))
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        
        ws.add_chart(chart1, "H6")
        
        # Risk Distribution Pie Chart
        risk_counts = df['Risk Category'].value_counts()
        
        chart2 = PieChart()
        chart2.title = "Risk Category Distribution"
        
        data = Reference(ws, min_col=2, min_row=27, max_row=27+len(risk_counts), max_col=2)
        cats = Reference(ws, min_col=1, min_row=28, max_row=27+len(risk_counts))
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
        
        ws.add_chart(chart2, "H20")
    
    def create_detailed_analysis_report(self):
        """Create detailed analysis report"""
        print("Creating detailed analysis report...")
        
        if self.processed_df is None:
            self.load_and_prepare_data()
        
        df = self.processed_df
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # 1. Policy Performance Analysis
        self._create_policy_performance_sheet(wb, df)
        
        # 2. Risk Analysis
        self._create_risk_analysis_sheet(wb, df)
        
        # 3. Customer Segmentation
        self._create_customer_segmentation_sheet(wb, df)
        
        # 4. Geographic Analysis
        self._create_geographic_analysis_sheet(wb, df)
        
        # 5. Reserve Monitoring
        self._create_reserve_monitoring_sheet(wb, df)
        
        return wb
    
    def _create_policy_performance_sheet(self, wb, df):
        """Create policy performance analysis sheet"""
        ws = wb.create_sheet("Policy Performance")
        
        # Title
        ws['A1'] = "POLICY PERFORMANCE ANALYSIS"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:H1')
        
        # Policy type performance with advanced formulas
        policy_analysis = df.groupby('Policy Type').agg({
            'Premium Amount': ['count', 'sum', 'mean', 'std'],
            'Previous Claims': ['sum', 'mean'],
            'Loss Ratio': ['mean', 'std'],
            'Customer Value': 'sum'
        }).round(2)
        
        policy_analysis.columns = ['Count', 'Total Premiums', 'Avg Premium', 'Std Premium',
                                 'Total Claims', 'Avg Claims', 'Avg Loss Ratio', 'Std Loss Ratio', 'Total Value']
        policy_analysis = policy_analysis.reset_index()
        
        # Add data to worksheet
        headers = ['Policy Type', 'Count', 'Total Premiums', 'Avg Premium', 'Std Premium',
                  'Total Claims', 'Avg Claims', 'Avg Loss Ratio', 'Std Loss Ratio', 'Total Value']
        
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=3, column=col_idx, value=header).font = Font(bold=True)
        
        for row_idx, (_, row) in enumerate(policy_analysis.iterrows(), start=4):
            for col_idx, value in enumerate([row['Policy Type'], row['Count'], 
                                          f"${row['Total Premiums']:,.2f}", f"${row['Avg Premium']:,.2f}",
                                          f"${row['Std Premium']:,.2f}", row['Total Claims'], 
                                          f"{row['Avg Claims']:.2f}", f"{row['Avg Loss Ratio']:.2%}",
                                          f"{row['Std Loss Ratio']:.2%}", f"${row['Total Value']:,.2f}"], start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Add conditional formatting formulas
        self._add_conditional_formatting(ws, df)
    
    def _create_risk_analysis_sheet(self, wb, df):
        """Create risk analysis sheet"""
        ws = wb.create_sheet("Risk Analysis")
        
        # Title
        ws['A1'] = "RISK ANALYSIS"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:H1')
        
        # Risk category analysis
        risk_analysis = df.groupby('Risk Category').agg({
            'Premium Amount': ['count', 'sum', 'mean'],
            'Previous Claims': ['sum', 'mean'],
            'Loss Ratio': 'mean',
            'Age': 'mean',
            'Health Score': 'mean',
            'Credit Score': 'mean'
        }).round(2)
        
        risk_analysis.columns = ['Count', 'Total Premiums', 'Avg Premium', 'Total Claims',
                               'Avg Claims', 'Avg Loss Ratio', 'Avg Age', 'Avg Health Score', 'Avg Credit Score']
        risk_analysis = risk_analysis.reset_index()
        
        # Add data to worksheet
        headers = ['Risk Category', 'Count', 'Total Premiums', 'Avg Premium', 'Total Claims',
                  'Avg Claims', 'Avg Loss Ratio', 'Avg Age', 'Avg Health Score', 'Avg Credit Score']
        
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=3, column=col_idx, value=header).font = Font(bold=True)
        
        for row_idx, (_, row) in enumerate(risk_analysis.iterrows(), start=4):
            for col_idx, value in enumerate([row['Risk Category'], row['Count'], 
                                          f"${row['Total Premiums']:,.2f}", f"${row['Avg Premium']:,.2f}",
                                          row['Total Claims'], f"{row['Avg Claims']:.2f}", 
                                          f"{row['Avg Loss Ratio']:.2%}", f"{row['Avg Age']:.1f}",
                                          f"{row['Avg Health Score']:.1f}", f"{row['Avg Credit Score']:.1f}"], start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
    
    def _create_customer_segmentation_sheet(self, wb, df):
        """Create customer segmentation sheet"""
        ws = wb.create_sheet("Customer Segmentation")
        
        # Title
        ws['A1'] = "CUSTOMER SEGMENTATION ANALYSIS"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:H1')
        
        # Customer segmentation by income and risk
        customer_segments = df.groupby(['Income Group', 'Risk Category']).agg({
            'Premium Amount': ['count', 'sum', 'mean'],
            'Customer Value': 'mean',
            'Previous Claims': 'mean',
            'Loss Ratio': 'mean',
            'Age': 'mean'
        }).round(2)
        
        customer_segments.columns = ['Count', 'Total Premiums', 'Avg Premium', 'Avg Customer Value',
                                   'Avg Claims', 'Avg Loss Ratio', 'Avg Age']
        customer_segments = customer_segments.reset_index()
        
        # Add data to worksheet
        headers = ['Income Group', 'Risk Category', 'Count', 'Total Premiums', 'Avg Premium',
                  'Avg Customer Value', 'Avg Claims', 'Avg Loss Ratio', 'Avg Age']
        
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=3, column=col_idx, value=header).font = Font(bold=True)
        
        for row_idx, (_, row) in enumerate(customer_segments.iterrows(), start=4):
            for col_idx, value in enumerate([row['Income Group'], row['Risk Category'], row['Count'],
                                          f"${row['Total Premiums']:,.2f}", f"${row['Avg Premium']:,.2f}",
                                          f"${row['Avg Customer Value']:,.2f}", f"{row['Avg Claims']:.2f}",
                                          f"{row['Avg Loss Ratio']:.2%}", f"{row['Avg Age']:.1f}"], start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
    
    def _create_geographic_analysis_sheet(self, wb, df):
        """Create geographic analysis sheet"""
        ws = wb.create_sheet("Geographic Analysis")
        
        # Title
        ws['A1'] = "GEOGRAPHIC ANALYSIS"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:H1')
        
        # Location performance analysis
        location_analysis = df.groupby('Location').agg({
            'Premium Amount': ['count', 'sum', 'mean'],
            'Previous Claims': ['sum', 'mean'],
            'Loss Ratio': 'mean',
            'Risk Score': 'mean',
            'Customer Value': 'sum'
        }).round(2)
        
        location_analysis.columns = ['Count', 'Total Premiums', 'Avg Premium', 'Total Claims',
                                   'Avg Claims', 'Avg Loss Ratio', 'Avg Risk Score', 'Total Value']
        location_analysis = location_analysis.reset_index()
        
        # Add data to worksheet
        headers = ['Location', 'Count', 'Total Premiums', 'Avg Premium', 'Total Claims',
                  'Avg Claims', 'Avg Loss Ratio', 'Avg Risk Score', 'Total Value']
        
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=3, column=col_idx, value=header).font = Font(bold=True)
        
        for row_idx, (_, row) in enumerate(location_analysis.iterrows(), start=4):
            for col_idx, value in enumerate([row['Location'], row['Count'], 
                                          f"${row['Total Premiums']:,.2f}", f"${row['Avg Premium']:,.2f}",
                                          row['Total Claims'], f"{row['Avg Claims']:.2f}", 
                                          f"{row['Avg Loss Ratio']:.2%}", f"{row['Avg Risk Score']:.1f}",
                                          f"${row['Total Value']:,.2f}"], start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
    
    def _create_reserve_monitoring_sheet(self, wb, df):
        """Create reserve monitoring sheet"""
        ws = wb.create_sheet("Reserve Monitoring")
        
        # Title
        ws['A1'] = "RESERVE MONITORING & CAPITAL ADEQUACY"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:H1')
        
        # Reserve calculations by risk category
        reserve_analysis = df.groupby('Risk Category').agg({
            'Premium Amount': ['sum', 'mean'],
            'Previous Claims': 'sum',
            'Loss Ratio': 'mean'
        }).round(2)
        
        reserve_analysis.columns = ['Total Premiums', 'Avg Premium', 'Total Claims', 'Avg Loss Ratio']
        reserve_analysis = reserve_analysis.reset_index()
        
        # Calculate reserves (simplified model)
        reserve_analysis['Required Reserves'] = reserve_analysis['Total Premiums'] * 0.15  # 15% of premiums
        reserve_analysis['Actual Claims'] = reserve_analysis['Total Claims'] * 1000  # Assuming $1000 per claim
        reserve_analysis['Reserve Surplus/Deficit'] = reserve_analysis['Required Reserves'] - reserve_analysis['Actual Claims']
        reserve_analysis['Reserve Ratio'] = reserve_analysis['Required Reserves'] / reserve_analysis['Actual Claims']
        
        # Add data to worksheet
        headers = ['Risk Category', 'Total Premiums', 'Avg Premium', 'Total Claims', 'Avg Loss Ratio',
                  'Required Reserves', 'Actual Claims', 'Reserve Surplus/Deficit', 'Reserve Ratio']
        
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=3, column=col_idx, value=header).font = Font(bold=True)
        
        for row_idx, (_, row) in enumerate(reserve_analysis.iterrows(), start=4):
            for col_idx, value in enumerate([row['Risk Category'], f"${row['Total Premiums']:,.2f}",
                                          f"${row['Avg Premium']:,.2f}", row['Total Claims'],
                                          f"{row['Avg Loss Ratio']:.2%}", f"${row['Required Reserves']:,.2f}",
                                          f"${row['Actual Claims']:,.2f}", f"${row['Reserve Surplus/Deficit']:,.2f}",
                                          f"{row['Reserve Ratio']:.2f}"], start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
    
    def _add_conditional_formatting(self, ws, df):
        """Add conditional formatting with Excel formulas"""
        # Add formulas for conditional formatting
        ws['K1'] = "CONDITIONAL FORMATTING FORMULAS"
        ws['K1'].font = Font(size=12, bold=True)
        
        # Loss ratio threshold formula
        ws['K3'] = "Loss Ratio Threshold:"
        ws['L3'] = "=IF(G4>0.7,\"High Risk\",\"Normal\")"
        
        # Premium variance formula
        ws['K4'] = "Premium Variance:"
        ws['L4'] = "=IF(D4>AVERAGE(D:D)*1.5,\"Above Average\",\"Normal\")"
        
        # Claims frequency formula
        ws['K5'] = "Claims Frequency:"
        ws['L5'] = "=IF(F4>AVERAGE(F:F)*1.2,\"High Claims\",\"Normal\")"
    
    def create_client_presentation_report(self):
        """Create client presentation report with advanced formatting"""
        print("Creating client presentation report...")
        
        if self.processed_df is None:
            self.load_and_prepare_data()
        
        df = self.processed_df
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Client Presentation"
        
        # Set up professional formatting
        self._setup_client_presentation_style(ws)
        
        # Title page
        ws['A1'] = "RISK & POLICY ANALYTICS DASHBOARD"
        ws['A1'].font = Font(size=20, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        ws.merge_cells('A1:H1')
        
        ws['A2'] = "Comprehensive Insurance Portfolio Analysis"
        ws['A2'].font = Font(size=14, italic=True, color="1F4E79")
        ws.merge_cells('A2:H2')
        
        ws['A3'] = f"Report Generated: {datetime.now().strftime('%B %d, %Y')}"
        ws['A3'].font = Font(size=10, color="666666")
        ws.merge_cells('A3:H3')
        
        # Executive Summary
        ws['A5'] = "EXECUTIVE SUMMARY"
        ws['A5'].font = Font(size=16, bold=True, color="1F4E79")
        ws.merge_cells('A5:H5')
        
        # Key insights
        total_policies = len(df)
        total_premiums = df['Premium Amount'].sum()
        avg_premium = df['Premium Amount'].mean()
        overall_loss_ratio = (df['Previous Claims'] * 1000).sum() / df['Premium Amount'].sum()
        
        insights = [
            f"• Portfolio contains {total_policies:,} active policies",
            f"• Total premium volume of ${total_premiums:,.2f}",
            f"• Average premium of ${avg_premium:,.2f} per policy",
            f"• Overall loss ratio of {overall_loss_ratio:.1%}",
            f"• {len(df[df['Risk Category'].isin(['High', 'Very High'])])} high-risk policies requiring attention"
        ]
        
        for idx, insight in enumerate(insights, start=7):
            ws[f'A{idx}'] = insight
            ws[f'A{idx}'].font = Font(size=11)
        
        # Recommendations
        ws['A13'] = "KEY RECOMMENDATIONS"
        ws['A13'].font = Font(size=16, bold=True, color="1F4E79")
        ws.merge_cells('A13:H13')
        
        recommendations = [
            "• Implement risk-based pricing adjustments for high-risk segments",
            "• Enhance underwriting criteria for policies with loss ratios > 80%",
            "• Develop targeted retention strategies for high-value customers",
            "• Consider premium adjustments for geographic areas with poor performance",
            "• Strengthen reserve requirements for high-risk policy categories"
        ]
        
        for idx, rec in enumerate(recommendations, start=15):
            ws[f'A{idx}'] = rec
            ws[f'A{idx}'].font = Font(size=11)
        
        # Add charts and tables
        self._add_client_presentation_charts(ws, df)
        
        return wb
    
    def _setup_client_presentation_style(self, ws):
        """Set up client presentation styling"""
        # Professional color scheme
        primary_color = "1F4E79"
        secondary_color = "D9E2F3"
        accent_color = "FFC000"
        
        # Set row heights
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[5].height = 25
        ws.row_dimensions[13].height = 25
        
        # Set column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 15
    
    def _add_client_presentation_charts(self, ws, df):
        """Add charts to client presentation"""
        # Premium distribution chart
        premium_ranges = pd.cut(df['Premium Amount'], bins=5)
        premium_dist = premium_ranges.value_counts().sort_index()
        
        chart1 = BarChart()
        chart1.title = "Premium Distribution"
        chart1.x_axis.title = "Premium Range"
        chart1.y_axis.title = "Number of Policies"
        
        # Risk category pie chart
        risk_counts = df['Risk Category'].value_counts()
        
        chart2 = PieChart()
        chart2.title = "Risk Category Distribution"
        
        # Add charts to worksheet
        ws.add_chart(chart1, "A20")
        ws.add_chart(chart2, "F20")
    
    def export_all_reports(self, output_path='04_Excel_Reports/'):
        """Export all Excel reports"""
        print(f"Exporting Excel reports to {output_path}...")
        
        # Create output directory if it doesn't exist
        import os
        os.makedirs(output_path, exist_ok=True)
        
        # Export executive summary
        exec_wb = self.create_executive_summary_report()
        exec_wb.save(f'{output_path}Executive_Summary_Report.xlsx')
        
        # Export detailed analysis
        detail_wb = self.create_detailed_analysis_report()
        detail_wb.save(f'{output_path}Detailed_Analysis_Report.xlsx')
        
        # Export client presentation
        client_wb = self.create_client_presentation_report()
        client_wb.save(f'{output_path}Client_Presentation_Report.xlsx')
        
        print("Excel reports exported successfully!")
        return True

def main():
    """Main execution function"""
    print("="*60)
    print("EXCEL AUTOMATION FOR RISK & POLICY ANALYTICS")
    print("="*60)
    
    # Initialize Excel automation
    excel_automation = ExcelAutomation('Insurance Premium Prediction Dataset.csv')
    
    # Load and prepare data
    excel_automation.load_and_prepare_data()
    
    # Export all reports
    excel_automation.export_all_reports()
    
    print("\n" + "="*60)
    print("EXCEL AUTOMATION COMPLETED!")
    print("="*60)
    
    return excel_automation

if __name__ == "__main__":
    excel_automation = main()
